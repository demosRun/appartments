import json
from openpyxl import load_workbook

# 加载 Excel 文件
wb = load_workbook('fy.xlsx')

# 选择工作表
ws = wb.active  # 或 wb['Sheet1']

obgTemp = {}
# 遍历内容
for row in ws.iter_rows(values_only=True):
    if (row[1] != 'Direct'):
        if (row[4] not in obgTemp):
            obgTemp[row[4]] = {
                "name": row[4],
                "Agoda Homes": [],
                "Agoda homes upc": [],
                "Agoda Hotel": [],
                "Booking": [],
                "Expedia": [],
                "Trip": [],
            }
        if (row[1] not in obgTemp[row[4]]):
            obgTemp[row[4]][row[1]] = []
        obgTemp[row[4]][row[1]].append(row[0])
        # obgTemp[row[4]][row[1]] = row[0]
        # obgTemp[row[4]][row[1] + 'Name'] = row[5]
print(obgTemp)
with open('room.json', 'w', encoding='utf-8') as f:
    json.dump(obgTemp, f, ensure_ascii=False, indent=2)